import io
from typing import Any, Dict, List

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, Image as PDFImage

import graphs



URBAN_GREEN_HEX = "2E7D32"  # Dark Green
URBAN_LIGHT_GREEN_HEX = "E8F5E9"  # Light Green
URBAN_GREEN_COLOR = colors.HexColor("#" + URBAN_GREEN_HEX)
URBAN_LIGHT_GREEN_COLOR = colors.HexColor("#" + URBAN_LIGHT_GREEN_HEX)

TRANSLATIONS = {
    "es": {
        "users": "Usuarios",
        "bicycles": "Bicicletas",
        "travels": "Viajes",
        "fines": "Multas",
        "fine_types": "Tipos de Multas",
        "user_fines": "Multas de Usuarios",
        "complaints": "Quejas",
        "maintenances": "Mantenimientos",
        "balance": "Saldo",
        "metric": "Métrica",
        "value": "Valor",
        "users_count": "Total Usuarios",
        "bicycles_count": "Total Bicicletas",
        "travels_count": "Total Viajes",
        "fines_count": "Total Multas",
        "complaints_count": "Total Quejas",
        "maintenance_count": "Total Mantenimientos",
        "complaints_by_status": "Quejas por estado",
        "complaints_by_type": "Quejas por tipo",
        "maintenance_by_entity": "Mantenimiento por entidad",
        "User": "Usuario",
        "Summary": "Resumen",

        "bicycle_usage": "Uso de Bicicletas",
        "station_demand": "Demanda Estaciones",
        "service_demand": "Demanda Servicio",
        "bicycle_demand": "Demanda Bicicletas",
        "daily_trips": "Viajes Diarios",

        "email": "Correo",
        "username": "Usuario",
        "subscription": "Suscripción",

        "n_user_name": "Nombre",
        "n_user_email": "Correo",
        "k_uid_user": "ID Usuario",
        "t_subscription_type": "Suscripción",
        "v_balance": "Saldo",
        "k_id_bicycle": "ID Bicicleta",
        "n_model": "Modelo",
        "f_last_update": "Última Actualización",
        "f_started_at": "Inicio",
        "f_ended_at": "Fin",
        "t_status": "Estado",
        "d_description": "Descripción",
        "f_date": "Fecha",
        "k_id_maintenance": "ID Mantenimiento",
        "t_entity_type": "Entidad",
        "t_maintance_type": "Tipo Mantenimiento",
        "t_triggered_by": "Iniciado Por",
        "v_cost": "Costo",
        "k_id_station": "ID Estación",
        "k_id_slot": "ID Slot",
        "n_station_name": "Estación",
        "n_latitude": "Latitud",
        "n_length": "Longitud",
        "t_type": "Tipo",
        "t_cctv_status": "CCTV",
        "k_id_city": "ID Ciudad",
        "k_id_fine": "ID Multa",
        "v_amount": "Monto",
        "k_id_complaints_and_claims": "ID Queja",
        "k_id_travel": "ID Viaje",
        "f_required_at": "Solicitado",
        "k_from_id_station": "Origen",
        "k_to_id_station": "Destino",
        "t_travel_type": "Tipo Viaje",
        "v_travel_cost": "Costo Viaje",
        "n_reason": "Razón",
        "t_state": "Estado Multa",
        "v_amount_snapshot": "Monto (Snapshot)",
        "f_assigned_at": "Asignada",
        "f_update_at": "Actualizada",
        "user_id": "ID Usuario",
        "status": "Estado",

        "date": "Fecha",
        "start_time": "Hora Inicio",
        "end_time": "Hora Fin",
        "bicycle": "Bicicleta",
        "origin_station": "Origen",
        "destination_station": "Destino",
        "travel_type": "Tipo Viaje",
        "description": "Descripción",
        "reason": "Razón",
        "amount_snapshot": "Monto",
        "assigned_at": "Asignada",
    },
    "en": {
        "users": "Users",
        "bicycles": "Bicycles",
        "travels": "Travels",
        "fines": "Fines",
        "fine_types": "Fine Types",
        "user_fines": "User Fines",
        "complaints": "Complaints",
        "maintenances": "Maintenances",
        "balance": "Balance",
        "metric": "Metric",
        "value": "Value",
        "users_count": "Total Users",
        "bicycles_count": "Total Bicycles",
        "travels_count": "Total Travels",
        "fines_count": "Total Fines",
        "complaints_count": "Total Complaints",
        "maintenance_count": "Total Maintenances",
        "complaints_by_status": "Complaints by Status",
        "complaints_by_type": "Complaints by Type",
        "maintenance_by_entity": "Maintenance by Entity",
        "User": "User",
        "n_user_email": "Email",
        "Summary": "Summary",

        "bicycle_usage": "Bicycle Usage",
        "station_demand": "Station Demand",
        "service_demand": "Service Demand",
        "bicycle_demand": "Bicycle Demand",
        "daily_trips": "Daily Trips",
    }
}


def _translate(key: str, lang: str) -> str:
    """Translate a key if found in the dictionary for the given language."""
    return TRANSLATIONS.get(lang, {}).get(key, key)


def _to_df(rows: List[Dict[str, Any]], lang: str = "en") -> pd.DataFrame:
    """Normalize list of dicts into a DataFrame, handling empty cases and translation."""
    if not rows:
        return pd.DataFrame({_translate("metric", lang): ["(No Data)"]})
    df = pd.DataFrame.from_records(rows)
    # Rename columns based on translation
    df.rename(columns=lambda x: _translate(x, lang), inplace=True)
    return df


def _summary_df_admin(overview: Dict[str, Any], lang: str = "en") -> pd.DataFrame:
    data = {
        "metric": [
            _translate("users_count", lang),
            _translate("bicycles_count", lang),
            _translate("travels_count", lang),
            _translate("fines_count", lang),
            _translate("complaints_count", lang),
            _translate("maintenance_count", lang),
        ],
        "value": [
            overview.get("users_count", 0),
            overview.get("bicycles_count", 0),
            overview.get("travels_count", 0),
            overview.get("fines_count", 0),
            overview.get("complaints_count", 0),
            overview.get("maintenance_count", 0),
        ],
    }
    df = pd.DataFrame(data)
    # Rename metric/value columns
    df.columns = [_translate("metric", lang), _translate("value", lang)]

    breakdown_rows = []
    for label, payload in (
        ("complaints_by_status", overview.get("complaints_by_status", {})),
        ("complaints_by_type", overview.get("complaints_by_type", {})),
        ("maintenance_by_entity", overview.get("maintenance_by_entity", {})),
    ):
        if payload:
            translated_label = _translate(label, lang)
            for key, value in payload.items():
                breakdown_rows.append({df.columns[0]: f"{translated_label}:{key}", df.columns[1]: value})
    if breakdown_rows:
        df = pd.concat([df, pd.DataFrame(breakdown_rows)], ignore_index=True)
    return df


def _summary_df_user(dashboard: Dict[str, Any], lang: str = "en") -> pd.DataFrame:
    data = {
        "metric": [
            _translate("travels_count", lang),
            _translate("complaints_count", lang),
            _translate("fines_count", lang),
            _translate("balance", lang),
        ],
        "value": [
            dashboard.get("travels_count", 0),
            dashboard.get("complaints_count", 0),
            dashboard.get("fines_count", 0),
            dashboard.get("balance", {}).get("balance")
            or dashboard.get("balance", {}).get("amount")
            or dashboard.get("balance", {}).get("v_amount"),
        ],
    }
    df = pd.DataFrame(data)
    df.columns = [_translate("metric", lang), _translate("value", lang)]
    return df




def _style_excel_sheet(ws):
    """Apply UrbanRide green styling to an Excel worksheet."""
    # Header Style
    header_fill = PatternFill(start_color=URBAN_GREEN_HEX, end_color=URBAN_GREEN_HEX, fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    # Apply to first row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(adjusted_width, 60)


def _add_graph_to_excel(writer, sheet_name, img_buffer):
    """Helper to add a graph to a specific Excel sheet."""
    if sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
    else:
        ws = writer.book.create_sheet(sheet_name)
    
    img = XLImage(img_buffer)
    ws.add_image(img, 'E2')


def build_admin_overview_excel_pandas(overview: Dict[str, Any]) -> bytes:
    """Create an Excel report using pandas with multiple sheets and styling."""
    lang = overview.get("lang", "en")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _summary_df_admin(overview, lang).to_excel(writer, sheet_name=_translate("Summary", lang), index=False)
        _to_df(overview.get("users", []), lang).to_excel(writer, sheet_name=_translate("users", lang), index=False)
        _to_df(overview.get("bicycles", []), lang).to_excel(writer, sheet_name=_translate("bicycles", lang), index=False)
        _to_df(overview.get("travels", []), lang).to_excel(writer, sheet_name=_translate("travels", lang), index=False)
        

        _to_df(overview.get("fine_types", []), lang).to_excel(writer, sheet_name=_translate("fine_types", lang), index=False)
        _to_df(overview.get("user_fines", []), lang).to_excel(writer, sheet_name=_translate("user_fines", lang), index=False)
        
        _to_df(overview.get("complaints", []), lang).to_excel(writer, sheet_name=_translate("complaints", lang), index=False)
        _to_df(overview.get("maintenances", []), lang).to_excel(writer, sheet_name=_translate("maintenances", lang), index=False)


        

        sheet_name = _translate("bicycle_usage", lang)
        _to_df(overview.get("bicycle_usage", []), lang).to_excel(writer, sheet_name=sheet_name, index=False)
        _add_graph_to_excel(writer, sheet_name, graphs.generate_bicycle_usage_chart(overview.get("bicycle_usage", []), lang))


        sheet_name = _translate("station_demand", lang)
        _to_df(overview.get("station_demand", {}).get("most_popular_origins", []), lang).to_excel(writer, sheet_name=sheet_name, index=False)
        _add_graph_to_excel(writer, sheet_name, graphs.generate_station_demand_chart(overview.get("station_demand", {}), lang))


        sheet_name = _translate("service_demand", lang)
        # Create a small DF for the stats
        service_data = overview.get("service_demand", {})
        pd.DataFrame([service_data]).to_excel(writer, sheet_name=sheet_name, index=False)
        _add_graph_to_excel(writer, sheet_name, graphs.generate_service_demand_chart(service_data, lang))


        sheet_name = _translate("bicycle_demand", lang)
        bike_demand_data = overview.get("bicycle_demand", {})
        pd.DataFrame([bike_demand_data]).to_excel(writer, sheet_name=sheet_name, index=False)
        _add_graph_to_excel(writer, sheet_name, graphs.generate_bicycle_type_demand_chart(bike_demand_data, lang))


        sheet_name = _translate("daily_trips", lang)
        _to_df(overview.get("daily_trips", []), lang).to_excel(writer, sheet_name=sheet_name, index=False)
        _add_graph_to_excel(writer, sheet_name, graphs.generate_daily_trips_chart(overview.get("daily_trips", []), lang))


        # Apply styles to all sheets
        for sheet_name in writer.sheets:
            _style_excel_sheet(writer.sheets[sheet_name])
            
    return output.getvalue()


def build_user_dashboard_excel_pandas(dashboard: Dict[str, Any]) -> bytes:
    """Create a user dashboard Excel report using pandas and styling."""
    lang = dashboard.get("lang", "en")
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        _summary_df_user(dashboard, lang).to_excel(writer, sheet_name=_translate("Summary", lang), index=False)
        _to_df([dashboard.get("user", {})], lang).to_excel(writer, sheet_name=_translate("User", lang), index=False)
        _to_df(dashboard.get("travels", []), lang).to_excel(writer, sheet_name=_translate("travels", lang), index=False)
        _to_df(dashboard.get("complaints", []), lang).to_excel(writer, sheet_name=_translate("complaints", lang), index=False)
        _to_df(dashboard.get("fines", []), lang).to_excel(writer, sheet_name=_translate("fines", lang), index=False)

        # Apply styles to all sheets
        for sheet_name in writer.sheets:
            _style_excel_sheet(writer.sheets[sheet_name])

    return output.getvalue()




def build_single_report_excel(data: List[Dict[str, Any]] | Dict[str, Any], title_key: str, lang: str = "es", graph_func=None) -> bytes:
    """Generic builder for a single report Excel file."""
    output = io.BytesIO()
    sheet_name = _translate(title_key, lang)
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df = pd.DataFrame()
        if isinstance(data, list):
            df = _to_df(data, lang)
        elif isinstance(data, dict):
            if "most_popular_origins" in data:
                 df = _to_df(data.get("most_popular_origins", []), lang)
            else:
                 df = pd.DataFrame([data])
        
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        if graph_func:
            _add_graph_to_excel(writer, sheet_name, graph_func(data, lang))
            
        _style_excel_sheet(writer.sheets[sheet_name])
        
    return output.getvalue()


def build_single_report_pdf(data: List[Dict[str, Any]] | Dict[str, Any], title_key: str, lang: str = "es", graph_func=None) -> bytes:
    """Generic builder for a single report PDF file."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    
    title_text = _translate(title_key, lang)
    title_style = styles["Title"]
    title_style.textColor = URBAN_GREEN_COLOR
    
    flow = [Paragraph(title_text, title_style), Spacer(1, 20)]
    
    if graph_func:
        _add_graph_to_pdf(flow, graph_func(data, lang), title_text)
        
    # Table
    df = pd.DataFrame()
    if isinstance(data, list):
        df = _to_df(data, lang)
    elif isinstance(data, dict):
        if "most_popular_origins" in data:
             df = _to_df(data.get("most_popular_origins", []), lang)
        else:
             df = pd.DataFrame([data])
             
    flow += _table_from_df(df, title_text)
    
    doc.build(flow)
    return buffer.getvalue()




def _table_from_df(df: pd.DataFrame, title: str, max_rows: int = 50) -> List[Any]:
    if df.empty:
        data = [["(no data)"]]
    else:
        # Styles
        styles = getSampleStyleSheet()
        
        # Header Style
        header_style = ParagraphStyle(
            'HeaderStyle',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=10,
            textColor=colors.whitesmoke,
            alignment=1, # Center
            leading=12,
        )
        
        # Row Style
        row_style = ParagraphStyle(
            'RowStyle',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=9,
            textColor=colors.black,
            alignment=1, # Center
            leading=11,
        )
        
        # Wrap Headers
        headers = [Paragraph(str(h), header_style) for h in df.columns.tolist()]
        
        # Wrap Rows
        rows = df.head(max_rows).fillna("").astype(str).values.tolist()
        wrapped_rows = []
        for row in rows:
            wrapped_row = []
            for cell in row:
                wrapped_row.append(Paragraph(str(cell), row_style))
            wrapped_rows.append(wrapped_row)
            
        data = [headers] + wrapped_rows
    
    table = Table(data, repeatRows=1)
    
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), URBAN_GREEN_COLOR),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, URBAN_GREEN_COLOR),
    ]
    
    # Alternating row colors
    if len(data) > 1:
        for i in range(1, len(data)):
            if i % 2 == 0:
                bg_color = URBAN_LIGHT_GREEN_COLOR
            else:
                bg_color = colors.white
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), bg_color))

    table.setStyle(TableStyle(style_cmds))
    
    title_style = getSampleStyleSheet()["Heading3"]
    title_style.textColor = URBAN_GREEN_COLOR
    title_paragraph = Paragraph(f"<b>{title}</b>", title_style)
    
    return [title_paragraph, Spacer(1, 6), table, Spacer(1, 12)]


def _add_graph_to_pdf(flow, img_buffer, title, width=400, height=240):
    """Helper to add graph to PDF flow."""
    styles = getSampleStyleSheet()
    title_style = styles["Heading3"]
    title_style.textColor = URBAN_GREEN_COLOR
    flow.append(Paragraph(f"<b>{title}</b>", title_style))
    flow.append(Spacer(1, 6))
    
    img = PDFImage(img_buffer, width=width, height=height)
    flow.append(img)
    flow.append(Spacer(1, 12))


def build_admin_overview_pdf(overview: Dict[str, Any]) -> bytes:
    lang = overview.get("lang", "en")
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    
    title_text = "UrbanRide - Admin Overview Report" if lang == "en" else "UrbanRide - Reporte General Admin"
    title_style = styles["Title"]
    title_style.textColor = URBAN_GREEN_COLOR
    
    flow = [Paragraph(title_text, title_style), Spacer(1, 20)]

    flow += _table_from_df(_summary_df_admin(overview, lang), _translate("Summary", lang))
    

    _add_graph_to_pdf(flow, graphs.generate_bicycle_usage_chart(overview.get("bicycle_usage", []), lang), _translate("bicycle_usage", lang))
    _add_graph_to_pdf(flow, graphs.generate_station_demand_chart(overview.get("station_demand", {}), lang), _translate("station_demand", lang))
    _add_graph_to_pdf(flow, graphs.generate_service_demand_chart(overview.get("service_demand", {}), lang), _translate("service_demand", lang))
    _add_graph_to_pdf(flow, graphs.generate_bicycle_type_demand_chart(overview.get("bicycle_demand", {}), lang), _translate("bicycle_demand", lang))
    _add_graph_to_pdf(flow, graphs.generate_daily_trips_chart(overview.get("daily_trips", []), lang), _translate("daily_trips", lang))
    
    # Tables
    flow += _table_from_df(_to_df(overview.get("users", []), lang), _translate("users", lang))
    flow += _table_from_df(_to_df(overview.get("bicycles", []), lang), _translate("bicycles", lang))
    flow += _table_from_df(_to_df(overview.get("travels", []), lang), _translate("travels", lang))
    

    flow += _table_from_df(_to_df(overview.get("fine_types", []), lang), _translate("fine_types", lang))
    flow += _table_from_df(_to_df(overview.get("user_fines", []), lang), _translate("user_fines", lang))
    
    flow += _table_from_df(_to_df(overview.get("complaints", []), lang), _translate("complaints", lang))
    flow += _table_from_df(_to_df(overview.get("maintenances", []), lang), _translate("maintenances", lang))

    doc.build(flow)
    return buffer.getvalue()


def build_user_dashboard_pdf(dashboard: Dict[str, Any]) -> bytes:
    lang = dashboard.get("lang", "en")
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    
    title_text = "UrbanRide - User Dashboard Report" if lang == "en" else "UrbanRide - Reporte Usuario"
    title_style = styles["Title"]
    title_style.textColor = URBAN_GREEN_COLOR
    
    flow = [Paragraph(title_text, title_style), Spacer(1, 20)]

    flow += _table_from_df(_summary_df_user(dashboard, lang), _translate("Summary", lang))
    flow += _table_from_df(_to_df([dashboard.get("user", {})], lang), _translate("User", lang))
    flow += _table_from_df(_to_df(dashboard.get("travels", []), lang), _translate("travels", lang))
    flow += _table_from_df(_to_df(dashboard.get("complaints", []), lang), _translate("complaints", lang))
    flow += _table_from_df(_to_df(dashboard.get("fines", []), lang), _translate("fines", lang))

    doc.build(flow)
    return buffer.getvalue()
