from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def _write_table(ws, title: str, rows: List[Dict[str, Any]]) -> None:
    ws.append([title])
    if not rows:
        ws.append(["(no data)"])
        ws.append([])
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    # auto width
    for i, header in enumerate(headers, start=1):
        max_len = max(len(str(header)), *(len(str(r.get(header))) for r in rows))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 60)
    ws.append([])


def build_admin_overview_excel(overview: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"

    ws.append(["UrbanRide - Admin Overview Report"])
    ws.append(["Users", overview.get("users_count", 0)])
    ws.append(["Bicycles", overview.get("bicycles_count", 0)])
    ws.append(["Reservations", overview.get("reservations_count", 0)])
    ws.append(["Fines", overview.get("fines_count", 0)])
    ws.append([])

    _write_table(ws, "Users", overview.get("users", []))
    _write_table(ws, "Bicycles", overview.get("bicycles", []))
    _write_table(ws, "Reservations", overview.get("reservations", []))
    _write_table(ws, "Fines", overview.get("fines", []))

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_user_dashboard_excel(dashboard: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "My Dashboard"

    user = dashboard.get("user", {})
    ws.append(["UrbanRide - User Dashboard Report"])
    ws.append(["User ID", user.get("id") or user.get("k_uid") or user.get("user_id")])
    ws.append(["Reservations", dashboard.get("reservations_count", 0)])
    ws.append(["Fines", dashboard.get("fines_count", 0)])
    balance = dashboard.get("balance", {})
    ws.append(["Balance", balance.get("balance") or balance.get("amount") or balance.get("v_amount")])
    ws.append([])

    _write_table(ws, "Reservations", dashboard.get("reservations", []))
    _write_table(ws, "Fines", dashboard.get("fines", []))

    from io import BytesIO
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
